import openpyxl
import requests
from lxml import etree
from openpyxl import Workbook

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36 Edg/106.0.1370.47'}
res = requests.get(
    'https://www.dotcpp.com/oj/problemset.php?page=1',
    headers=headers)  # 获取网页,page可更改

html = etree.HTML(res.text)

"""
https://www.dotcpp.com/robots.txt:

User-agent: *
Sitemap: https://www.dotcpp.com/sitemap.xml
Disallow: /oj/contestrank.php?*
Disallow: /oj/contestrank-oi.php?*
Disallow: /oj/conteststatistics.php?*
Disallow: /oj/showsource.php?*
Disallow: /oj/status.php?*
Disallow: /oj/status.html?*
Disallow: /oj/mail.php?*
Disallow: /oj/loginpage.php
Disallow: /oj/registerpage.php
Disallow: /oj/loginpage_cpn.php
Disallow: /oj/intro.php
Disallow: /oj/business.php
Disallow: /oj/contactus.php
Disallow: /job/*
Disallow: /vipstudy_suanfa/*
Disallow: /?*
"""

xlsx_list = []  # 用于存储xlsx文件中的数据

tr_number = res.text.count('/tr')   # 统计tr标签的数量
for i in range(1, tr_number - 1):
    all_data = []
    number = html.xpath(
        f'//*[@id="problemset"]/tbody/tr[{i}]/td[2]/span')  # 编号
    question = html.xpath(
        f'//*[@id="problemset"]/tbody/tr[{i}]/td[3]/span/a/h3')  # 题目
    question_type = html.xpath(
        f'//*[@id="problemset"]/tbody/tr[{i}]/td[4]/a')  # 类型
    difficulty = html.xpath(
        f'//*[@id="problemset"]/tbody/tr[{i}]/td[5]/a')  # 难度
    resolve_and_commit = html.xpath(
        f'//*[@id="problemset"]/tbody/tr[{i}]/td[6]/span')  # 解决/提交

    all_data.append(number[0].text.strip())
    all_data.append(question[0].text.strip())
    all_data.append(question_type[0].text.strip())
    all_data.append(difficulty[0].text.strip())
    all_data.append(resolve_and_commit[0].text.strip())

    xlsx_list.append(all_data)  # 将每一行的数据添加到xlsx_list中

# 保存为excel

try:
    workbook = openpyxl.load_workbook('dotcpp.xlsx')    # 打开excel
except FileNotFoundError:   # 文件不存在
    workbook = Workbook()   # 创建工作簿
    sheet = workbook.active     # 获取当前活跃的sheet,默认是第一个sheet
    sheet.title = 'dotcpp'  # 修改sheet名
    workbook['dotcpp'].append(['编号', '题目', '类型', '难度', '解决/提交'])    # 添加表头
    sheet.column_dimensions['B'].width = 35     # 设置列宽
    sheet.column_dimensions['C'].width = 9
    sheet.column_dimensions['E'].width = 13
    workbook.save('dotcpp.xlsx')    # 保存文件
    workbook = openpyxl.load_workbook('dotcpp.xlsx')    # 重新加载文件

for row in xlsx_list:
    workbook['dotcpp'].append(row)  # 添加数据
workbook.save(filename='dotcpp.xlsx')   # 保存文件
